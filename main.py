import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# ### in terminal do: pip install openpyxl
exel_file = xl.load_workbook('EnglishData.xlsx')
excel_sheet = exel_file['Sheet1']

def find_is_string_exist(user_input2, list1):
    for word_from_list in list1:
        if user_input2 == word_from_list:
            return 1
    return 0


hebrew = ""
english = ""
for row in range(1, excel_sheet.max_row + 1):
    cell = excel_sheet.cell(row, 2)
    hebrew = cell.value
    cell = excel_sheet.cell(row, 1)
    english = cell.value
    hebrew = hebrew.replace(' ; ', ', ')
    hebrew_list = hebrew.split(', ')
    print(f"What is the translation for the word: {english}")
    i = 1
    while i < 4:
        print(f"remaining tries number {i}")
        user_input = input("> ")
        result = find_is_string_exist(user_input, hebrew_list)
        if result == 0:
            print("incorrect")
            i = i + 1
        else:
            hebrew_list.remove(user_input)
            if len(hebrew_list) == 0:
                break
            else:
                print(f"you have {len(hebrew_list)} more meaning to this word")
    if i < 4:
        print("Good Job")
    else:
        print("Nice try you get it Right next time")

